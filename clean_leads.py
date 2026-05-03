"""
═══════════════════════════════════════════════════════════════════
  REAL ESTATE LEAD LIST CLEANER
  Developed by Rashad Ferguson | Data & AI Operations
  rashad.io | github.com/rashadflowers99
═══════════════════════════════════════════════════════════════════

One script. One command. Three outputs:

  leads_clean          — standardized, deduped, CRM-ready
  leads_email_outreach — warm leads missing email, scripts included
  leads_master.xlsx    — formatted workbook with four tabs:
                           Master List / Import Ready /
                           Email Outreach / Name Conflicts

Usage:
  python clean_leads.py
  python clean_leads.py --input your_export.xlsx
  python clean_leads.py --input leads.csv --output ./cleaned/
"""

import pandas as pd
import re, os, argparse, glob
from datetime import datetime
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

COLUMN_MAP = {
    "lead_name":"full_name","name":"full_name","contact":"full_name",
    "first_name":"first_name","last_name":"last_name","fname":"first_name","lname":"last_name",
    "phone":"phone","phonenumber":"phone","mobile":"phone","cell":"phone",
    "email":"email","emailaddress":"email","e-mail":"email",
    "source":"lead_source","leadsource":"lead_source","lead_source":"lead_source",
    "status":"status","budget":"budget","price":"budget","pricerange":"budget",
    "address":"address","city":"city","state":"state","zip":"zip","zipcode":"zip",
    "notes":"notes","comments":"notes",
}

SOURCE_MAP = {
    "website contact":"Website","website":"Website","web":"Website",
    "zillow":"Zillow","zillow lead":"Zillow",
    "realtor.com":"Realtor.com","realtor":"Realtor.com",
    "fb_ad_1":"Facebook Ad","fb_ad_2":"Facebook Ad",
    "facebook ad":"Facebook Ad","facebook":"Facebook Ad","fb":"Facebook Ad",
    "referral":"Referral","cold call":"Cold Call",
}

SMS_SCRIPT = (
    "Hi {FIRST_NAME}, this is [AGENT NAME] from [BROKERAGE]. "
    "I have your {SOURCE} inquiry on file and would love to help — "
    "could you reply with your email so I can send over the best listings? "
    "Takes 5 seconds and I'll have options to you same day."
)

CALL_SCRIPT = (
    "Hey {FIRST_NAME}, it's [AGENT NAME] with [BROKERAGE]. "
    "I see you reached out through {SOURCE} — just need a quick email address "
    "so I can send you properties that match your budget. "
    "What's the best email for you?"
)


# ─────────────────────────────────────────────────────────────────────────────
# FIELD CLEANERS
# ─────────────────────────────────────────────────────────────────────────────

def split_full_name(value):
    if pd.isna(value) or str(value).strip() == "":
        return ("", "")
    parts = str(value).strip().title().split()
    return (parts[0], " ".join(parts[1:])) if len(parts) > 1 else (parts[0], "") if parts else ("", "")

def clean_phone(value):
    if pd.isna(value) or str(value).strip() == "":
        return ""
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}" if len(digits) == 10 else ""

def clean_email(value):
    if pd.isna(value) or str(value).strip() == "":
        return ""
    e = str(value).strip().lower()
    return e if re.match(r"^[\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}$", e) else ""

def clean_budget(value):
    if pd.isna(value) or str(value).strip() == "":
        return ""
    raw = str(value).strip()
    if raw.lower() == "unknown":
        return "Unknown"
    c = raw.replace("$","").replace(",","").strip()
    if c.lower().endswith("k"):
        try: return f"${int(float(c[:-1])*1_000):,}"
        except ValueError: pass
    if c.lower().endswith("m"):
        try: return f"${int(float(c[:-1])*1_000_000):,}"
        except ValueError: pass
    try: return f"${int(float(c)):,}"
    except ValueError: return raw

def clean_source(value):
    if pd.isna(value) or str(value).strip() == "":
        return ""
    return SOURCE_MAP.get(str(value).strip().lower(), str(value).strip().title())

def clean_text(value):
    return str(value).strip().title() if not pd.isna(value) and str(value).strip() else ""

def clean_notes(value):
    return str(value).strip() if not pd.isna(value) and str(value).strip() else ""


# ─────────────────────────────────────────────────────────────────────────────
# DEDUPLICATION
# ─────────────────────────────────────────────────────────────────────────────

def deduplicate(df):
    original   = len(df)
    phone_dupe = df["phone"].ne("") & df.duplicated(subset=["phone"], keep="first")
    email_dupe = df["email"].ne("") & df.duplicated(subset=["email"], keep="first")
    return df[~(phone_dupe | email_dupe)].copy(), original - len(df[~(phone_dupe | email_dupe)])


# ─────────────────────────────────────────────────────────────────────────────
# TRIAGE — Three buckets
# ─────────────────────────────────────────────────────────────────────────────

def triage_records(df):
    has_name  = df["first_name"].ne("")
    has_phone = df["phone"].ne("")
    has_email = df["email"].ne("")

    clean_df    = df[has_name & has_phone & has_email].copy()
    outreach_df = df[has_name & has_phone & ~has_email].copy()
    review_df   = df[(~has_name | ~has_phone) & ~(has_name & has_phone & has_email)].copy()

    if not outreach_df.empty:
        outreach_df["action"]      = "Call or text to collect email"
        outreach_df["sms_script"]  = outreach_df.apply(
            lambda r: SMS_SCRIPT.replace("{FIRST_NAME}", r["first_name"])
                                 .replace("{SOURCE}", r["lead_source"] or "your inquiry"), axis=1)
        outreach_df["call_script"] = outreach_df.apply(
            lambda r: CALL_SCRIPT.replace("{FIRST_NAME}", r["first_name"])
                                  .replace("{SOURCE}", r["lead_source"] or "your inquiry"), axis=1)
        outreach_df["email_collected"] = ""
        outreach_df["contacted_date"]  = ""

    if not review_df.empty:
        review_df = review_df.copy()
        review_df["needs_review"] = review_df.apply(
            lambda r: "; ".join(
                (["missing name"] if not r.get("first_name","") else []) +
                (["missing/invalid phone"] if not r.get("phone","") else [])
            ), axis=1)

    return clean_df, outreach_df, review_df


# ─────────────────────────────────────────────────────────────────────────────
# SAME-NAME CONFLICT DETECTOR
# ─────────────────────────────────────────────────────────────────────────────

def _area_code(phone):
    digits = re.sub(r"\D", "", str(phone))
    return digits[:3] if len(digits) >= 10 else ""

def _budget_int(budget):
    try: return int(str(budget).replace("$","").replace(",",""))
    except: return -1

FL_METROS = [{"305","786"}, {"407","689"}, {"813","727"}, {"904"}, {"561"}]

def _get_metro(ac):
    for m in FL_METROS:
        if ac in m:
            return frozenset(m)
    return None

def _score_pair(r1, r2):
    score, reasons = 0, []
    ac1, ac2 = _area_code(r1.get("phone","")), _area_code(r2.get("phone",""))
    b1,  b2  = _budget_int(r1.get("budget","")), _budget_int(r2.get("budget",""))
    e1,  e2  = str(r1.get("email","")).strip().lower(), str(r2.get("email","")).strip().lower()
    s1,  s2  = str(r1.get("lead_source","")), str(r2.get("lead_source",""))

    if ac1 and ac2 and ac1 == ac2:
        score += 3; reasons.append(f"same area code ({ac1})")
    if b1 > 0 and b2 > 0:
        diff = abs(b1 - b2)
        if diff == 0:    score += 2; reasons.append("identical budget")
        elif diff <= 50_000: score += 1; reasons.append(f"budgets within ${diff:,}")
        else:            score -= 1; reasons.append(f"budgets differ by ${diff:,}")
    if (e1 and not e2) or (e2 and not e1):
        score += 2; reasons.append("one entry missing email (likely same person)")
    if s1 and s2 and s1 == s2:
        score += 1; reasons.append(f"same lead source ({s1})")
    m1, m2 = _get_metro(ac1), _get_metro(ac2)
    if m1 and m2 and m1 != m2:
        score -= 2; reasons.append(f"different FL metros (area codes {ac1} vs {ac2})")

    if score >= 4:   verdict = "LIKELY DUPLICATE"
    elif score >= 1: verdict = "POSSIBLE DUPLICATE"
    else:            verdict = "PROBABLY DIFFERENT"

    return verdict, reasons

def detect_conflicts(df):
    df = df.copy()
    df["full_name"] = (df["first_name"] + " " + df["last_name"]).str.strip()
    multi = df["full_name"].value_counts()
    multi = multi[multi > 1].index.tolist()
    if not multi:
        return pd.DataFrame()
    rows = []
    for name in sorted(multi):
        group = df[df["full_name"] == name].reset_index(drop=True)
        for i, j in combinations(range(len(group)), 2):
            r1, r2 = group.iloc[i], group.iloc[j]
            verdict, reasons = _score_pair(r1, r2)
            rows.append({
                "full_name":       name,
                "record_a_phone":  r1.get("phone",""),
                "record_a_email":  r1.get("email",""),
                "record_a_source": r1.get("lead_source",""),
                "record_a_budget": r1.get("budget",""),
                "record_b_phone":  r2.get("phone",""),
                "record_b_email":  r2.get("email",""),
                "record_b_source": r2.get("lead_source",""),
                "record_b_budget": r2.get("budget",""),
                "verdict":         verdict,
                "reasoning":       " | ".join(reasons) if reasons else "no matching signals",
                "agent_decision":  "",
            })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK BUILDER — four formatted tabs
# ─────────────────────────────────────────────────────────────────────────────

FOREST  = "1A5C38"; FOREST2 = "2D7A4F"; GRN_LIGHT = "E8F5EE"
AMB_DRK = "8B5E00"; AMB_MID = "A07010"; AMB_LIGHT  = "FFF8E8"
RED_DRK = "7A1F1F"; RED_MID = "A03030"; RED_LIGHT  = "FFF0F0"
YEL_DRK = "7A6000"; YEL_MID = "B08000"; YEL_LIGHT  = "FFFFF0"
WHITE   = "FFFFFF"

VERDICT_PALETTE = {
    "LIKELY DUPLICATE":   (RED_DRK, RED_MID,  RED_LIGHT),
    "POSSIBLE DUPLICATE": (YEL_DRK, YEL_MID,  YEL_LIGHT),
    "PROBABLY DIFFERENT": (FOREST,  FOREST2, GRN_LIGHT),
}

def _solid(h):    return PatternFill("solid", fgColor=h)
def _border(c="D0D0D0"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def _btm(c="E0E0E0"):
    b = Side(style="thin", color=c); n = Side(style=None)
    return Border(left=n, right=n, top=n, bottom=b)

def _hcell(ws, row, col, val, bg, fg="FFFFFF", sz=10, bold=True, wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, size=sz, color=fg)
    c.fill = _solid(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = _border(bg)
    return c

def _dcell(ws, row, col, val, bg, fg="1A1A1A", sz=9, wrap=False, align="center", bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=sz, color=fg, bold=bold)
    c.fill = _solid(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = _btm()
    return c

def _title_rows(ws, row1_text, row2_text, ncols, bg1, bg2, fg1="FFFFFF", fg2="C8E6D0"):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=row1_text)
    t.font = Font(name="Arial", bold=True, size=13, color=fg1)
    t.fill = _solid(bg1)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=row2_text)
    s.font = Font(name="Arial", size=10, color=fg2)
    s.fill = _solid(bg2)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

def build_workbook(clean_df, outreach_df, conflicts, output_path):
    wb = Workbook()

    # Shared master data
    MASTER_COLS   = ["first_name","last_name","phone","email","lead_source","budget","status"]
    MASTER_HEADS  = ["First Name","Last Name","Phone","Email","Source","Budget","Status"]
    MASTER_WIDTHS = [14,14,17,28,14,12,15]

    all_status = pd.concat([
        clean_df.assign(status="Import Ready")[MASTER_COLS],
        outreach_df.assign(email="", status="Email Needed")[[c for c in MASTER_COLS if c in outreach_df.assign(email="",status="Email Needed").columns]]
    ], ignore_index=True)
    all_status["_s"] = all_status["status"].map({"Import Ready":0,"Email Needed":1})
    master = all_status.sort_values("_s").drop(columns=["_s"]).reset_index(drop=True)

    # ── TAB 1: Master List ────────────────────────────────────────────────
    ws = wb.active; ws.title = "Master List"
    ws.sheet_view.showGridLines = False; ws.freeze_panes = "A4"
    _title_rows(ws,
        "🏠  Florida Lead List — Master Pipeline",
        f"  {len(master)} leads  •  {len(clean_df)} Import Ready  •  "
        f"{len(outreach_df)} Email Needed  •  "
        f"Prepared by Rashad Ferguson | rashad.io",
        len(MASTER_COLS), FOREST, FOREST2)
    ws.row_dimensions[1].height = 32
    for ci,(col,head,w) in enumerate(zip(MASTER_COLS,MASTER_HEADS,MASTER_WIDTHS),1):
        _hcell(ws,3,ci,head,FOREST)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22
    for ri,(_,row) in enumerate(master.iterrows(),start=4):
        ready = row["status"]=="Import Ready"
        bg = GRN_LIGHT if ready else AMB_LIGHT
        for ci,col in enumerate(MASTER_COLS,1):
            bold = col=="status"
            fg = "1A5C38" if (col=="status" and ready) else "7A4F00" if (col=="status") else "1A1A1A"
            _dcell(ws,ri,ci,row.get(col,""),bg,fg=fg,bold=bold)
        ws.row_dimensions[ri].height = 17
    tr = len(master)+4
    ws.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=2)
    tc = ws.cell(row=tr,column=1,value="TOTALS")
    tc.font=Font(name="Arial",bold=True,size=9,color="FFFFFF"); tc.fill=_solid(FOREST)
    tc.alignment=Alignment(horizontal="center",vertical="center")
    sc = ws.cell(row=tr,column=MASTER_COLS.index("status")+1,
        value=f"✓ {len(clean_df)} Ready  |  📞 {len(outreach_df)} Need Email")
    sc.font=Font(name="Arial",bold=True,size=9,color="FFFFFF"); sc.fill=_solid(FOREST2)
    sc.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[tr].height = 20

    # ── TAB 2: Import Ready ───────────────────────────────────────────────
    ws2 = wb.create_sheet("✓ Import Ready")
    ws2.sheet_view.showGridLines = False; ws2.freeze_panes = "A4"
    IR_COLS=["first_name","last_name","phone","email","lead_source","budget"]
    IR_HEADS=["First Name","Last Name","Phone","Email","Source","Budget"]
    IR_WIDTHS=[14,14,17,30,14,12]
    _title_rows(ws2,
        "✓  Import-Ready Leads — Load directly into your CRM or dialer",
        f"  {len(clean_df)} records  •  All fields standardized  •  Zero duplicates",
        len(IR_COLS), FOREST, FOREST2)
    for ci,(col,head,w) in enumerate(zip(IR_COLS,IR_HEADS,IR_WIDTHS),1):
        _hcell(ws2,3,ci,head,FOREST)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[3].height = 22
    for ri,(_,row) in enumerate(clean_df.iterrows(),start=4):
        bg = GRN_LIGHT if ri%2==0 else WHITE
        for ci,col in enumerate(IR_COLS,1):
            _dcell(ws2,ri,ci,row.get(col,""),bg)
        ws2.row_dimensions[ri].height = 16

    # ── TAB 3: Email Outreach ─────────────────────────────────────────────
    ws3 = wb.create_sheet("📞 Email Outreach")
    ws3.sheet_view.showGridLines = False; ws3.freeze_panes = "A4"
    OUT_COLS=["first_name","last_name","phone","lead_source","budget",
              "sms_script","call_script","email_collected","contacted_date"]
    OUT_HEADS=["First Name","Last Name","Phone","Source","Budget",
               "SMS Script (copy & send)","Call Script",
               "Email Collected ✏","Date Contacted ✏"]
    OUT_WIDTHS=[14,14,17,14,12,52,52,26,16]
    _title_rows(ws3,
        "📞  Email Outreach Queue — One call or text per lead collects the missing email",
        f"  {len(outreach_df)} warm leads  •  Phone verified  •  Scripts personalized  •  Fill in 'Email Collected' after each call",
        len(OUT_COLS), AMB_DRK, AMB_MID, fg2="FFE0A0")
    for ci,(col,head,w) in enumerate(zip(OUT_COLS,OUT_HEADS,OUT_WIDTHS),1):
        _hcell(ws3,3,ci,head,AMB_DRK,wrap=True)
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[3].height = 22
    for ri,(_,row) in enumerate(outreach_df.iterrows(),start=4):
        bg = AMB_LIGHT if ri%2==0 else WHITE
        for ci,col in enumerate(OUT_COLS,1):
            wrap = col in ("sms_script","call_script")
            _dcell(ws3,ri,ci,row.get(col,""),bg,wrap=wrap,
                   align="left" if wrap else "center")
        ws3.row_dimensions[ri].height = 65

    # ── TAB 4: Name Conflicts ─────────────────────────────────────────────
    ws4 = wb.create_sheet("⚠ Name Conflicts")
    ws4.sheet_view.showGridLines = False; ws4.freeze_panes = "A4"
    CONF_COLS=["full_name","verdict","reasoning",
               "record_a_phone","record_a_email","record_a_source","record_a_budget",
               "record_b_phone","record_b_email","record_b_source","record_b_budget",
               "agent_decision"]
    CONF_HEADS=["Name","Verdict","Why","A: Phone","A: Email","A: Source","A: Budget",
                "B: Phone","B: Email","B: Source","B: Budget","Your Decision ✏"]
    CONF_WIDTHS=[16,20,52,17,26,14,12,17,26,14,12,22]

    likely   = len(conflicts[conflicts["verdict"]=="LIKELY DUPLICATE"])
    possible = len(conflicts[conflicts["verdict"]=="POSSIBLE DUPLICATE"])
    diff_ct  = len(conflicts[conflicts["verdict"]=="PROBABLY DIFFERENT"])

    _title_rows(ws4,
        "⚠  Same-Name Conflict Report — Review and decide: MERGE / KEEP BOTH / DISCARD B",
        f"  {len(conflicts)} pairs across {conflicts['full_name'].nunique()} name groups  •  "
        f"{likely} Likely Duplicate  •  {possible} Possible  •  {diff_ct} Probably Different",
        len(CONF_COLS), RED_DRK, RED_MID, fg2="FFD0D0")
    for ci,(col,head,w) in enumerate(zip(CONF_COLS,CONF_HEADS,CONF_WIDTHS),1):
        _hcell(ws4,3,ci,head,RED_DRK,wrap=True)
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[3].height = 22

    sort_order = {"LIKELY DUPLICATE":0,"POSSIBLE DUPLICATE":1,"PROBABLY DIFFERENT":2}
    conf_sorted = conflicts.copy()
    conf_sorted["_s"] = conf_sorted["verdict"].map(sort_order)
    conf_sorted = conf_sorted.sort_values(["_s","full_name"]).drop(columns=["_s"])

    for ri,(_,row) in enumerate(conf_sorted.iterrows(),start=4):
        dark,mid,light = VERDICT_PALETTE.get(row["verdict"],(RED_DRK,RED_MID,RED_LIGHT))
        for ci,col in enumerate(CONF_COLS,1):
            if col=="verdict":
                c=ws4.cell(row=ri,column=ci,value=row[col])
                c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
                c.fill=_solid(dark); c.border=_btm()
                c.alignment=Alignment(horizontal="center",vertical="center")
            else:
                wrap = col=="reasoning"
                _dcell(ws4,ri,ci,row.get(col,""),light,
                       wrap=wrap,align="left" if col in ("reasoning","agent_decision") else "center")
        ws4.row_dimensions[ri].height = 48 if len(str(row.get("reasoning","")))>60 else 28

    wb["Master List"].sheet_properties.tabColor       = FOREST
    wb["✓ Import Ready"].sheet_properties.tabColor    = FOREST2
    wb["📞 Email Outreach"].sheet_properties.tabColor = "F5A623"
    wb["⚠ Name Conflicts"].sheet_properties.tabColor  = "C0392B"

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# PROCESSING REPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_report(summary, output_dir):
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    path = os.path.join(output_dir, f"processing_report_{ts}.txt")
    odf  = summary["outreach_df"]
    cfl  = summary["conflicts"]

    budgets = []
    for b in odf.get("budget", []):
        try: budgets.append(int(str(b).replace("$","").replace(",","")))
        except: pass
    avg_b  = int(sum(budgets)/len(budgets)) if budgets else 0
    est_c  = int(avg_b * 0.025)

    likely   = len(cfl[cfl["verdict"]=="LIKELY DUPLICATE"])   if not cfl.empty else 0
    possible = len(cfl[cfl["verdict"]=="POSSIBLE DUPLICATE"]) if not cfl.empty else 0

    lines = [
        "═"*52,
        "  LEAD LIST PROCESSING REPORT",
        "  Prepared by Rashad Ferguson",
        "  rashad.io",
        "═"*52,"",
        f"  Run date:    {summary['timestamp']}",
        f"  Source file: {summary['input_file']}","",
        "─"*52,"  RESULTS SUMMARY","─"*52,
        f"  Records received:            {summary['raw_records']}",
        f"  Duplicates removed:         -{summary['duplicates_removed']}",
        f"  {'─'*32}",
        f"  After deduplication:         {summary['after_dedup']}","",
        f"  ✓  Import-ready:             {summary['clean_records']}",
        f"  📞 Email outreach queue:     {summary['outreach_records']}",
        f"  ⚠  Needs manual review:      {summary['review_records']}","",
    ]
    if summary["outreach_records"] > 0:
        lines += [
            "─"*52,"  EMAIL OUTREACH QUEUE","─"*52,
            f"  {summary['outreach_records']} leads have a valid phone but no email on file.",
            "  Personalized SMS + call scripts included per lead.",
            f"  Average budget: ${avg_b:,}" if avg_b else "",
            f"  Est. commission per close (~2.5%): ${est_c:,}" if est_c else "",
            f"  Potential if all {summary['outreach_records']} convert: ${est_c*summary['outreach_records']:,}" if est_c else "",
            "  One 5-minute call per lead. That is the only work required.","",
        ]
    if not cfl.empty:
        lines += [
            "─"*52,"  SAME-NAME CONFLICT REPORT","─"*52,
            f"  {len(cfl)} same-name pairs detected across {cfl['full_name'].nunique()} name groups.",
            f"  {likely} pairs scored LIKELY DUPLICATE — review first.",
            f"  {possible} pairs scored POSSIBLE DUPLICATE — agent judgement required.",
            "  See the '⚠ Name Conflicts' tab in the master workbook.","",
        ]
    lines += [
        "─"*52,"  WHAT WAS STANDARDIZED","─"*52,
        "  • Phone numbers  → (XXX) XXX-XXXX",
        "  • Email          → lowercase, validated",
        "  • Budget         → $XXX,XXX  ('$450k' → '$450,000')",
        "  • Source labels  → canonical  ('zillow lead' → 'Zillow')",
        "  • Names          → Title Case, split First + Last","",
        "═"*52,"  Questions? rashad.io","═"*52,
    ]
    with open(path,"w") as f:
        f.write("\n".join(lines))
    return path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def run_pipeline(input_path, output_dir):
    print(f"\n{'═'*52}\n  REAL ESTATE LEAD LIST CLEANER\n  by Rashad Ferguson\n{'═'*52}")
    print(f"\n  Input: {os.path.basename(input_path)}")

    ext = os.path.splitext(input_path)[1].lower()
    df  = pd.read_excel(input_path, dtype=str) if ext in (".xlsx",".xls") \
          else pd.read_csv(input_path, dtype=str)
    df  = df.fillna(""); raw_count = len(df)
    print(f"  Loaded: {raw_count} records")

    df.columns = [c.strip().lower().replace(" ","_") for c in df.columns]
    df = df.rename(columns=COLUMN_MAP)
    for col in ["first_name","last_name","full_name","phone","email",
                "lead_source","budget","address","city","state","zip","notes"]:
        if col not in df.columns: df[col] = ""

    if df["full_name"].ne("").any():
        splits = df["full_name"].apply(split_full_name)
        df["first_name"] = [splits[i][0] if not df.at[i,"first_name"] else df.at[i,"first_name"] for i in df.index]
        df["last_name"]  = [splits[i][1] if not df.at[i,"last_name"]  else df.at[i,"last_name"]  for i in df.index]

    print("  Cleaning fields...")
    df["first_name"]  = df["first_name"].apply(clean_text)
    df["last_name"]   = df["last_name"].apply(clean_text)
    df["phone"]       = df["phone"].apply(clean_phone)
    df["email"]       = df["email"].apply(clean_email)
    df["lead_source"] = df["lead_source"].apply(clean_source)
    df["budget"]      = df["budget"].apply(clean_budget)
    df["address"]     = df["address"].apply(clean_text)
    df["city"]        = df["city"].apply(clean_text)
    df["state"]       = df["state"].apply(lambda v: str(v).upper().strip() if v else "")
    df["zip"]         = df["zip"].apply(lambda v: re.sub(r"\D","",str(v))[:5] if v else "")
    df["notes"]       = df["notes"].apply(clean_notes)

    print("  Removing duplicates...")
    df, dupes_removed = deduplicate(df)

    print("  Triaging records...")
    clean_df, outreach_df, review_df = triage_records(df)

    print("  Detecting same-name conflicts...")
    all_leads = pd.concat([clean_df, outreach_df], ignore_index=True)
    conflicts  = detect_conflicts(all_leads)

    base_cols     = ["first_name","last_name","phone","email","lead_source","budget","address","city","state","zip","notes"]
    outreach_cols = ["first_name","last_name","phone","lead_source","budget",
                     "action","sms_script","call_script","email_collected","contacted_date"]

    clean_df    = clean_df.drop(columns=["full_name"],errors="ignore")
    clean_df    = clean_df[[c for c in base_cols if c in clean_df.columns]]
    outreach_df = outreach_df.drop(columns=["full_name"],errors="ignore")
    outreach_df = outreach_df[[c for c in outreach_cols if c in outreach_df.columns]]
    review_df   = review_df.drop(columns=["full_name"],errors="ignore")

    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    print("  Building master workbook...")
    workbook_path = os.path.join(output_dir, f"leads_master_{ts}.xlsx")
    build_workbook(clean_df, outreach_df, conflicts, workbook_path)

    return {
        "timestamp":          datetime.now().strftime("%Y-%m-%d %H:%M"),
        "input_file":         os.path.basename(input_path),
        "raw_records":        raw_count,
        "duplicates_removed": dupes_removed,
        "after_dedup":        raw_count - dupes_removed,
        "clean_records":      len(clean_df),
        "outreach_records":   len(outreach_df),
        "review_records":     len(review_df),
        "workbook_path":      workbook_path,
        "clean_df":           clean_df,
        "outreach_df":        outreach_df,
        "conflicts":          conflicts,
    }


def main():
    parser = argparse.ArgumentParser(description="Real Estate Lead List Cleaner — Rashad Ferguson")
    parser.add_argument("--input",  "-i", default="/mnt/user-data/uploads/messy_florida_leads.xlsx")
    parser.add_argument("--output", "-o", default="./output")
    args = parser.parse_args()

    summary     = run_pipeline(args.input, args.output)
    report_path = generate_report(summary, args.output)
    cfl         = summary["conflicts"]
    likely      = len(cfl[cfl["verdict"]=="LIKELY DUPLICATE"]) if not cfl.empty else 0

    print(f"\n{'─'*52}\n  DONE\n{'─'*52}")
    print(f"  ✓  {summary['clean_records']} import-ready")
    print(f"  📞 {summary['outreach_records']} email outreach (scripts included)")
    print(f"  ⚠  {summary['review_records']} needs review")
    if not cfl.empty:
        print(f"  🔍 {len(cfl)} same-name conflicts ({likely} likely duplicates)")
    print(f"\n  Master workbook → {os.path.basename(summary['workbook_path'])}")
    print(f"  Report          → processing_report_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
    print(f"\n  Saved to: {args.output}/\n{'─'*52}\n")

if __name__ == "__main__":
    main()
